"""
Admin export endpoint — orders and withdrawals to CSV/Excel/PDF.
"""
import io
import csv
from datetime import datetime
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from app.db.session import get_db
from app.models.order import Order
from app.models.affiliate import AffiliateWithdrawal
from app.models.user import User
from app.api.deps import get_current_admin
from app.core.config import settings

router = APIRouter(prefix="/admin/export", tags=["admin"])


@router.get("/orders")
async def export_orders(
    fmt: str = Query("csv", regex="^(csv|excel|pdf)$"),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    result = await db.execute(
        select(Order)
        .where(Order.deleted_at.is_(None))
        .options(selectinload(Order.items), selectinload(Order.user))
        .order_by(Order.created_at.desc())
    )
    orders = result.scalars().all()

    rows = []
    for o in orders:
        for item in o.items:
            snap = item.product_snapshot or {}
            product_url = f"{settings.FRONTEND_URL}/?product={item.product_id}" if item.product_id else "N/A"
            rows.append({
                "Order ID": str(o.id)[:8].upper(),
                "Date": o.created_at.strftime("%Y-%m-%d %H:%M"),
                "Customer": o.user.full_name if o.user else (o.contact_info or {}).get("name", "Guest"),
                "Email": o.user.email if o.user else (o.contact_info or {}).get("email", ""),
                "Phone": (o.contact_info or {}).get("phone", ""),
                "Product": snap.get("name", ""),
                "Product Link": product_url,
                "Qty": item.quantity,
                "Unit Price (MWK)": item.unit_price,
                "Total (MWK)": item.unit_price * item.quantity,
                "Payment": o.payment_method,
                "Status": o.status,
                "Affiliate": item.affiliate_id or "",
            })

    if fmt == "csv":
        return _csv_response(rows, "orders")
    elif fmt == "excel":
        return _excel_response(rows, "orders")
    else:
        return _pdf_response(rows, "Orders Report", "orders")


@router.get("/withdrawals")
async def export_withdrawals(
    fmt: str = Query("csv", regex="^(csv|excel|pdf)$"),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    result = await db.execute(
        select(AffiliateWithdrawal, User.full_name, User.email)
        .join(User, AffiliateWithdrawal.user_id == User.id)
        .where(AffiliateWithdrawal.deleted_at.is_(None))
        .order_by(AffiliateWithdrawal.created_at.desc())
    )
    rows_raw = result.all()
    rows = [
        {
            "ID": str(r.AffiliateWithdrawal.id)[:8].upper(),
            "Date": r.AffiliateWithdrawal.created_at.strftime("%Y-%m-%d %H:%M"),
            "Affiliate Name": r.full_name,
            "Email": r.email,
            "Amount (MWK)": r.AffiliateWithdrawal.amount,
            "Method": r.AffiliateWithdrawal.method,
            "Status": r.AffiliateWithdrawal.status,
            "Reviewed At": r.AffiliateWithdrawal.reviewed_at.strftime("%Y-%m-%d %H:%M") if r.AffiliateWithdrawal.reviewed_at else "",
        }
        for r in rows_raw
    ]

    if fmt == "csv":
        return _csv_response(rows, "withdrawals")
    elif fmt == "excel":
        return _excel_response(rows, "withdrawals")
    else:
        return _pdf_response(rows, "Withdrawals Report", "withdrawals")


def _csv_response(rows: list, name: str):
    if not rows:
        rows = [{"No data": ""}]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)
    filename = f"pamsika_{name}_{datetime.now().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _excel_response(rows: list, name: str):
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = name.capitalize()
        if rows:
            ws.append(list(rows[0].keys()))
            for row in rows:
                ws.append(list(row.values()))
            # Style header
            from openpyxl.styles import Font, PatternFill
            gold = "C8A84B"
            for cell in ws[1]:
                cell.font = Font(bold=True, color="000000")
                cell.fill = PatternFill("solid", fgColor=gold)
            # Auto width
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"pamsika_{name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ImportError:
        return _csv_response(rows, name)


def _pdf_response(rows: list, title: str, name: str):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
        styles = getSampleStyleSheet()
        elements = []
        elements.append(Paragraph(f"Pa_mSikA — {title}", styles["Title"]))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
        elements.append(Spacer(1, 12))

        if rows:
            headers = list(rows[0].keys())
            data = [headers] + [[str(r.get(h, "")) for h in headers] for r in rows]
            col_width = (landscape(A4)[0] - 40) / len(headers)
            t = Table(data, colWidths=[col_width] * len(headers))
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C8A84B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F0E8")]),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("PADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(t)

        doc.build(elements)
        output.seek(0)
        filename = f"pamsika_{name}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ImportError:
        return _csv_response(rows, name)
